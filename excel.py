from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


def salvar_resultados_excel(dados, pasta_destino="resultados"):
    """Gera um relatório Excel formatado com os resultados do saneamento."""

    # Se existir um arquivo chamado 'resultados', remove para criar a pasta
    if os.path.isfile(pasta_destino):
        os.remove(pasta_destino)

    os.makedirs(pasta_destino, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Saneamento D2D"

    # Cabeçalhos
    headers = [
        "Ação Fiscal / Chave",
        "Nota Fiscal",
        "Situação Imposto",
        "Status Final",
    ]
    ws.append(headers)

    # Estilo do Cabeçalho (Azul com texto Branco)
    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Cores (Verde para Liberada, Vermelho para Pendente)
    fill_verde = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    font_verde = Font(color="006100", bold=True)

    fill_vermelho = PatternFill(
        start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
    )
    font_vermelho = Font(color="9C0006", bold=True)

    # Preenchimento dos dados
    for item in dados:
        row = [
            item.get("acao_fiscal", ""),
            item.get("nota", ""),
            item.get("imposto", ""),
            item.get("situacao", "").upper(),
        ]
        ws.append(row)

        row_idx = ws.max_row
        status_cell = ws.cell(row=row_idx, column=4)

        if "LIBERADA" in str(status_cell.value):
            status_cell.fill = fill_verde
            status_cell.font = font_verde
        elif "PENDENTE" in str(status_cell.value) or "ERRO" in str(
            status_cell.value
        ):
            status_cell.fill = fill_vermelho
            status_cell.font = font_vermelho

        status_cell.alignment = Alignment(horizontal="center")

    # Ajusta largura das colunas
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # Nomeia o arquivo com Data e Hora
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    caminho_arquivo = os.path.join(
        pasta_destino, f"Saneamento_{timestamp}.xlsx"
    )

    wb.save(caminho_arquivo)
    return caminho_arquivo